from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction, models
from decimal import Decimal
from datetime import timedelta, date, datetime
from django.utils import timezone
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.db.models import Sum, Count

from apps.ventas.models import Pedido, DetallePedido
from apps.productos.models import Producto
from apps.clientes.models import Cliente
from apps.documentos.models import DocumentoVenta, DetalleDocumento, Pago

from apps.ventas.forms import (
    PedidoForm, TipoDocumentoForm, BoletaForm, 
    FacturaForm, CheckoutForm
)
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction
from django.db.models import F
from django.utils import timezone
from decimal import Decimal
from django.db.models import Q
from apps.productos.models import Producto

# VISTAS DEL CARRITO DE CLIENTE

@login_required
def cliente_add_to_cart(request, producto_id):
    if request.user.rol != 'Cliente':
        return redirect('usuarios:dashboard')

    producto = get_object_or_404(Producto, id=producto_id)
    
    try:
        quantity = int(request.POST.get('quantity', 1))
        if quantity <= 0:
            quantity = 1
    except ValueError:
        quantity = 1

    cart = request.session.get('cart', {})
    
    if str(producto_id) in cart:
        new_quantity = cart[str(producto_id)] + quantity
    else:
        new_quantity = quantity

    if new_quantity > producto.stock:
        messages.error(request, f'Stock insuficiente para {producto.nombre}. Solo quedan {producto.stock} unidades.')
        return redirect('usuarios:dashboard') 

    cart[str(producto_id)] = new_quantity
    request.session['cart'] = cart
    
    messages.success(request, f'"{producto.nombre}" añadido al carrito.')
    return redirect('usuarios:dashboard')


@login_required
def cliente_view_cart(request):
    if request.user.rol != 'Cliente':
        return redirect('usuarios:dashboard')

    cart = request.session.get('cart', {})
    if not cart:
        return render(request, 'ventas/cart.html', {'cart_items': [], 'total_carrito': 0})

    product_ids = cart.keys()
    productos_en_db = Producto.objects.filter(id__in=product_ids)
    
    cart_items = []
    total_carrito = Decimal('0.00')

    for producto in productos_en_db:
        cantidad = cart[str(producto.id)]
        subtotal = producto.precio_unitario * cantidad
        total_carrito += subtotal
        
        cart_items.append({
            'producto': producto,
            'cantidad': cantidad,
            'subtotal': subtotal,
        })

    context = {
        'cart_items': cart_items,
        'total_carrito': total_carrito,
    }
    return render(request, 'ventas/cart.html', context)


@login_required
def cliente_remove_from_cart(request, producto_id):
    if request.user.rol != 'Cliente':
        return redirect('usuarios:dashboard')

    cart = request.session.get('cart', {})
    
    if str(producto_id) in cart:
        del cart[str(producto_id)]
        request.session['cart'] = cart
        messages.success(request, 'Producto eliminado del carrito.')

    return redirect('ventas:cliente_view_cart')


@login_required
def cliente_checkout(request):
    """
    Muestra el formulario de checkout y procesa la compra.
    Ahora redirige al cliente al documento (Boleta/Factura) tras la compra.
    """
    # Lógica de verificación de cliente y carrito
    if request.user.rol != 'Cliente':
        return redirect('usuarios:dashboard')

    cart = request.session.get('cart', {})
    if not cart:
        messages.warning(request, 'Tu carrito está vacío.')
        return redirect('ventas:cliente_view_cart')

    try:
        # Intenta obtener el perfil del cliente
        cliente_actual = request.user.perfil_cliente
    except Cliente.DoesNotExist:
        messages.error(request, '¡Debes completar tu perfil antes de poder comprar!')
        return redirect('clientes:completar_perfil')

    # Lógica de cálculo del carrito
    product_ids = cart.keys()
    productos_en_db = Producto.objects.filter(id__in=product_ids)
    
    cart_items = []
    total_carrito = Decimal('0.00')
    for producto in productos_en_db:
        cantidad = cart.get(str(producto.id), 0)
        if cantidad > 0:
            subtotal = producto.precio_unitario * cantidad
            total_carrito += subtotal
            cart_items.append({
                'producto': producto,
                'cantidad': cantidad,
                'subtotal': subtotal,
            })


    # Procesamiento del POST
    if request.method == 'POST':
        form = CheckoutForm(request.POST, instance=cliente_actual)
        tipo_documento = request.POST.get('tipo_documento', 'Boleta')

        if tipo_documento == 'Factura':
            form.fields['giro'].required = True

        if form.is_valid():
            cliente_actual_guardado = form.save() 
            medio_de_pago = form.cleaned_data['medio_de_pago']
            
            try:
                with transaction.atomic():
                    # 1. Crear Pedido
                    nuevo_pedido = Pedido.objects.create(
                        cliente=cliente_actual_guardado,
                        usuario=request.user, 
                        total=total_carrito,
                        estado='Pendiente'
                    )
                    
                    # 2. Crear DetallePedido y descontar stock
                    for item in cart_items:
                        producto = item['producto']
                        cantidad = item['cantidad']
                        if producto.stock < cantidad:
                            raise Exception(f"Stock insuficiente para {producto.nombre}.")
                        DetallePedido.objects.create(
                            pedido=nuevo_pedido,
                            producto=producto,
                            cantidad=cantidad,
                            precio_unitario_venta=producto.precio_unitario 
                        )
                        producto.stock -= cantidad
                        producto.save()
                    
                    # 3. Calcular totales para Documento
                    total_bruto = total_carrito.quantize(Decimal('0.00'))
                    neto_calculado = (total_bruto / Decimal('1.19')).quantize(Decimal('0.00'))
                    iva_calculado = total_bruto - neto_calculado
                    fecha_actual = timezone.now()

                    # 4. Crear DocumentoVenta
                    documento = DocumentoVenta.objects.create(
                        pedido=nuevo_pedido,
                        tipo_documento=tipo_documento,
                        cliente=cliente_actual_guardado,
                        vendedor=request.user,
                        neto=neto_calculado,
                        iva=iva_calculado,
                        total=total_bruto,
                        fecha_emision=fecha_actual,
                        fecha_vencimiento=fecha_actual.date(), 
                        estado='Emitida', 
                        medio_de_pago=medio_de_pago,
                        razon_social=cliente_actual_guardado.razon_social,
                        rut=cliente_actual_guardado.rut,
                        giro=cliente_actual_guardado.giro,
                        direccion=cliente_actual_guardado.direccion
                    )
                    
                    # 5. Crear DetalleDocumento
                    for item in cart_items:
                        DetalleDocumento.objects.create(
                            documento=documento,
                            producto=item['producto'],
                            cantidad=item['cantidad'],
                            precio_unitario_venta=item['producto'].precio_unitario
                        )
                        
                    # 6. Crear Pago
                    Pago.objects.create(
                        documento=documento,
                        monto_pagado=total_bruto,
                        metodo_pago=medio_de_pago,
                        referencia="Pago E-Commerce"
                    )
                    
                    # 7. Limpiar carrito y Redirigir (¡PUNTO CORREGIDO!)
                    del request.session['cart']
                    messages.success(request, f'¡Compra realizada con éxito! {tipo_documento} #{documento.folio} ha sido generada y pagada. Puedes verla a continuación.')
                    
                    # Redirección al detalle del documento recién creado
                    return redirect('documentos:detalle_documento', documento_id=documento.id)

            except Exception as e:
                print(f"Error al procesar el pedido: {e}") 
                messages.error(request, f'Error al procesar el pedido: {str(e)}')
        
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')

    else: 
        form = CheckoutForm(instance=cliente_actual)

    context = {
        'form': form,
        'cart_items': cart_items,
        'total_carrito': total_carrito,
    }
    return render(request, 'ventas/checkout.html', context)

# VISTAS DE PEDIDOS (VENDEDOR / ADMIN)
# (Lógica de IVA e


@login_required
def listar_pedidos(request):
    # Inicializar la consulta base
    pedidos = Pedido.objects.select_related('cliente', 'usuario').prefetch_related('detalles').all()
    
    # --- FILTRO DE SEGURIDAD PARA CLIENTES ---
    if request.user.rol == 'Cliente':
        try:
            # Si es cliente, solo ve sus pedidos
            perfil_cliente = request.user.perfil_cliente
            pedidos = pedidos.filter(cliente=perfil_cliente)
        except Cliente.DoesNotExist:
            pedidos = pedidos.none()
            messages.warning(request, "⚠️ Debes completar tu perfil para ver tus pedidos.")

    filtro_usuario = request.GET.get('usuario')
    filtro_cliente = request.GET.get('cliente')

    if filtro_usuario:
        pedidos = pedidos.filter(usuario__username__icontains=filtro_usuario)
    if filtro_cliente:
        pedidos = pedidos.filter(cliente__razon_social__icontains=filtro_cliente)
    
    # Ordenar y finalizar
    pedidos = pedidos.order_by('-fecha_creacion')

    context = {
        'pedidos': pedidos,
        'estados': Pedido.ESTADOS_PEDIDO 
    }
    return render(request, 'ventas/listar_pedidos.html', context)


@login_required
def crear_pedido_inicial(request):
    if request.method == 'POST':
        form = PedidoForm(request.POST)
        if form.is_valid():
            pedido = form.save(commit=False)
            pedido.estado = 'Borrador'
            pedido.usuario = request.user  
            pedido.save()
            messages.success(request, f'Pedido #{pedido.id} creado. Ahora define el tipo de documento.')
            return redirect('ventas:crear_pedido_datos', pedido_id=pedido.id)
    else:
        form = PedidoForm()
    return render(request, 'ventas/crear_pedido_inicial.html', {'form': form})


@login_required
def crear_pedido_datos(request, pedido_id):
    """
    Paso 2: Vendedor añade datos del Documento (Factura/Boleta)
    (Cálculo de IVA "hacia atrás" y lógica Contado/Crédito)
    """
    pedido = get_object_or_404(Pedido, id=pedido_id)

    if request.method == 'POST':
        post = request.POST.copy()
        modalidad = post.get('modalidad_pago', 'ahora')
        tipo_doc = post.get('tipo_documento')

        if modalidad == 'ahora' and not post.get('fecha_emision'):
            post['fecha_emision'] = timezone.localdate().isoformat()
            post['fecha_vencimiento'] = timezone.localdate().isoformat()
        
        pedido_form = PedidoForm(post, instance=pedido)
        tipo_form = TipoDocumentoForm(post) 
        boleta_form = BoletaForm(post)
        factura_form = FacturaForm(post)

        if not (pedido_form.is_valid() and tipo_form.is_valid()):
            messages.error(request, 'Revisa los datos del pedido y el tipo de documento.')
        
        else:
            pedido_form.save()
            tipo_doc = tipo_form.cleaned_data['tipo_documento']
            
            doc = DocumentoVenta()
            doc.pedido = pedido
            doc.cliente = pedido.cliente
            doc.tipo_documento = tipo_doc
            doc.vendedor = request.user 

            # Lógica de IVA
            total_bruto = sum((dp.subtotal for dp in pedido.detalles.all()), Decimal('0'))
            neto_calculado = (total_bruto / Decimal('1.19')).quantize(Decimal('0.00'))
            iva_calculado = total_bruto - neto_calculado

            doc.neto = neto_calculado
            doc.iva = iva_calculado
            doc.total = total_bruto

            if tipo_doc == 'Boleta':
                if boleta_form.is_valid():
                    doc.medio_de_pago = boleta_form.cleaned_data['medio_de_pago']
                    doc.fecha_emision = timezone.now()
                    doc.fecha_vencimiento = timezone.now().date()
                    doc.estado = 'Pagada' 
                else:
                    messages.error(request, 'Revisa los datos de la boleta.')

            elif tipo_doc == 'Factura':
                if not factura_form.is_valid():
                    messages.error(request, 'Revisa los datos de la factura.')
                else:
                    doc.razon_social = factura_form.cleaned_data['razon_social']
                    doc.rut = factura_form.cleaned_data['rut']
                    doc.giro = factura_form.cleaned_data['giro']
                    doc.direccion = factura_form.cleaned_data['direccion']
                    doc.ciudad = factura_form.cleaned_data['ciudad']
                    doc.comuna = factura_form.cleaned_data['comuna']
                    doc.medio_de_pago = factura_form.cleaned_data['medio_de_pago']
                    
                    fecha_emision_user = factura_form.cleaned_data.get('fecha_emision')
                    fecha_venc_user = factura_form.cleaned_data.get('fecha_vencimiento')

                    if modalidad == 'plazos':
                        doc.fecha_emision = fecha_emision_user or date.today()
                        if fecha_venc_user:
                            doc.fecha_vencimiento = fecha_venc_user
                        elif post.get('dias_plazo'):
                            try:
                                dias = int(post.get('dias_plazo'))
                                doc.fecha_vencimiento = doc.fecha_emision + timedelta(days=dias)
                            except (ValueError, TypeError):
                                doc.fecha_vencimiento = None
                        doc.estado = 'Emitida'
                    
                    else: 
                        doc.fecha_emision = fecha_emision_user or timezone.localdate()
                        doc.fecha_vencimiento = doc.fecha_emision 
                        doc.estado = 'Pagada' 

            form_valido_para_guardar = (
                (tipo_doc == 'Boleta' and boleta_form.is_valid()) or
                (tipo_doc == 'Factura' and factura_form.is_valid())
            )

            if form_valido_para_guardar:
                doc.save() 
                DetalleDocumento.objects.filter(documento=doc).delete()
                
                for d in pedido.detalles.all():
                    DetalleDocumento.objects.create(
                        documento=doc,
                        producto=d.producto,
                        cantidad=d.cantidad,
                        precio_unitario_venta=d.precio_unitario_venta, 
                        costo_unitario_venta=getattr(d.producto, 'costo_unitario', Decimal('0')),
                    )
                
                # Si es "Pagada", crear el registro de Pago
                if doc.estado == 'Pagada':
                     Pago.objects.create(
                        documento=doc,
                        monto_pagado=doc.total,
                        metodo_pago=doc.medio_de_pago,
                        referencia="Pago Vendedor Contado"
                    )

                messages.success(request, f'{tipo_doc} #{doc.folio} creada. Ahora agrega los productos.')
                return redirect('ventas:agregar_productos_pedido', pedido_id=pedido.id)

    # GET (o si el formulario POST falló)
    pedido_form = PedidoForm(instance=pedido)
    tipo_form = TipoDocumentoForm()
    boleta_form = BoletaForm()
    factura_form = FacturaForm()

    return render(request, 'ventas/crear_pedido_datos.html', {
        'pedido_form': pedido_form,
        'tipo_form': tipo_form,
        'boleta_form': boleta_form,
        'factura_form': factura_form,
        'pedido': pedido,
    })


@login_required
def agregar_productos_pedido(request, pedido_id):
    """Paso 3: Vendedor agrega productos al Pedido en borrador y confirma."""

    # Obtener pedido y documento
    pedido = get_object_or_404(
        Pedido.objects.select_related('cliente', 'usuario', 'documentoventa'),
        id=pedido_id
    )

    if not hasattr(pedido, 'documentoventa') or not pedido.documentoventa:
        messages.error(request, "Este pedido no tiene un documento asociado. Vuelva al Paso 1.")
        return redirect('ventas:listar_pedidos')

    documento = pedido.documentoventa

    # Datos para la vista
    productos = Producto.objects.filter(activo=True).order_by('nombre')
    detalles = DetallePedido.objects.filter(pedido=pedido).select_related('producto')

    carrito = [{
        'producto_id': d.producto.id,
        'nombre': d.producto.nombre,
        'cantidad': d.cantidad,
        'precio': d.precio_unitario_venta,
        'subtotal': d.subtotal,
    } for d in detalles]

    mensaje_error = None


 
    if request.method == 'POST':

        # -------- CONFIRMAR PEDIDO --------
        if 'confirmar_pedido' in request.POST:
            if not detalles.exists():
                messages.error(request, "⚠️ No puede confirmar un pedido sin productos.")
                return redirect('ventas:agregar_productos_pedido', pedido_id=pedido.id)

            try:
                with transaction.atomic():

                    # 1. Descontar stock
                    for detalle in detalles:
                        producto = detalle.producto
                        if producto.stock < detalle.cantidad:
                            raise Exception(f"Stock insuficiente para {producto.nombre}. Disponible: {producto.stock}")
                        producto.stock -= detalle.cantidad
                        producto.save()

                    # 2. Cambiar estado según documento
                    pedido.estado = 'Procesando' if documento.estado == 'Pagada' else 'Pendiente'
                    pedido.save()

                    # 3. Actualizar detalles del documento
                    DetalleDocumento.objects.filter(documento=documento).delete()

                    for detalle in detalles:
                        DetalleDocumento.objects.create(
                            documento=documento,
                            producto=detalle.producto,
                            cantidad=detalle.cantidad,
                            precio_unitario_venta=detalle.precio_unitario_venta,
                            subtotal=detalle.subtotal,
                            costo_unitario_venta=detalle.producto.costo_unitario
                        )

                    messages.success(request, f" Pedido #{pedido.id} confirmado y stock descontado.")
                    return redirect('ventas:detalle_pedido', pedido_id=pedido.id)

            except Exception as e:
                messages.error(request, f" Error al confirmar: {str(e)}")
                return redirect('ventas:agregar_productos_pedido', pedido_id=pedido.id)

        # -------- AGREGAR PRODUCTO --------
        else:
            producto_id = request.POST.get('producto_id')
            cantidad = request.POST.get('cantidad')

            if not producto_id or not cantidad:
                mensaje_error = "⚠️ Debe seleccionar un producto y especificar la cantidad."
            else:
                try:
                    cantidad = int(cantidad)
                    producto = get_object_or_404(Producto, id=producto_id)

                    with transaction.atomic():
                        detalle_existente = DetallePedido.objects.filter(pedido=pedido, producto=producto).first()

                        if detalle_existente:
                            nueva_cantidad = detalle_existente.cantidad + cantidad

                            if producto.stock < nueva_cantidad:
                                mensaje_error = f"⚠️ Stock insuficiente. Ya tiene {detalle_existente.cantidad} unidades."
                            else:
                                detalle_existente.cantidad = nueva_cantidad
                                detalle_existente.save()
                                messages.success(request, f" Cantidad actualizada: {producto.nombre}")
                        else:
                            DetallePedido.objects.create(
                                pedido=pedido,
                                producto=producto,
                                cantidad=cantidad,
                                precio_unitario_venta=producto.precio_unitario
                            )
                            messages.success(request, f" Producto agregado: {producto.nombre}")

                        actualizar_totales_documento(pedido, documento)
                        return redirect('ventas:agregar_productos_pedido', pedido_id=pedido.id)

                except ValueError:
                    mensaje_error = " La cantidad debe ser un número válido."
                except Exception as e:
                    mensaje_error = f" Error al agregar el producto: {str(e)}"


    #       RENDERIZAR VISTA
    return render(request, 'ventas/agregar_productos_pedido.html', {
        'pedido': pedido,
        'productos': productos,
        'carrito': carrito,
        'mensaje_error': mensaje_error
    })



def actualizar_totales_documento(pedido, documento):
    """Función auxiliar para recalcular totales (IVA INCLUIDO)"""
    detalles = DetallePedido.objects.filter(pedido=pedido)
    
    total_bruto = sum(d.subtotal for d in detalles)
    neto_calculado = (total_bruto / Decimal('1.19')).quantize(Decimal('0.00'))
    iva_calculado = total_bruto - neto_calculado
    
    documento.neto = neto_calculado
    documento.iva = iva_calculado
    documento.total = total_bruto
    documento.save()
    
    pedido.total = total_bruto 
    pedido.save()


@login_required
def eliminar_producto_carrito(request, pedido_id, producto_id):
    """Elimina un producto del pedido (del vendedor)"""
    pedido = get_object_or_404(Pedido, id=pedido_id)
    producto = get_object_or_404(Producto, id=producto_id)
    
    try:
        with transaction.atomic():
            detalle = DetallePedido.objects.filter(pedido=pedido, producto=producto).first()
            
            if detalle:
                detalle.delete() 
                
                if hasattr(pedido, 'documentoventa') and pedido.documentoventa:
                    actualizar_totales_documento(pedido, pedido.documentoventa)
                
                messages.success(request, f"Producto eliminado: {producto.nombre}")
            else:
                messages.warning(request, " El producto no está en el pedido.")
                
    except Exception as e:
        messages.error(request, f" Error al eliminar el producto: {str(e)}")
    
    return redirect('ventas:agregar_productos_pedido', pedido_id=pedido_id)


@login_required
def detalle_pedido(request, pedido_id):
    pedido = get_object_or_404(
        Pedido.objects.select_related('cliente', 'usuario')
        .prefetch_related('detalles__producto'),
        id=pedido_id
    )
    detalles = pedido.detalles.all()
    try:
        documento = DocumentoVenta.objects.get(pedido=pedido)
    except DocumentoVenta.DoesNotExist:
        documento = None

    context = {
        'pedido': pedido,
        'detalles': detalles,
        'documento': documento,
    }
    return render(request, 'ventas/detalle_pedido.html', context)





@login_required
def confirmar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)

    print(f"[confirmar_pedido] inicio pedido_id={pedido_id}, estado actual={pedido.estado}")

    if pedido.estado != 'Pendiente':
        messages.warning(request, 'Este pedido ya fue confirmado o procesado.')
        return redirect('ventas:detalle_pedido', pedido_id=pedido.id)

    detalles = DetallePedido.objects.filter(pedido=pedido).select_related('producto')

    if not detalles.exists():
        messages.error(request, 'No se puede confirmar un pedido sin productos.')
        return redirect('ventas:detalle_pedido', pedido_id=pedido.id)

    # ids de productos
    product_ids = [d.producto.id for d in detalles if d.producto]
    print("[confirmar_pedido] productos en pedido:", product_ids)

    try:
        with transaction.atomic():
            productos_locked = Producto.objects.select_for_update().filter(id__in=product_ids)
            prod_map = {p.id: p for p in productos_locked}

            # 1) Verificar stock suficiente usando datos bloqueados 
            insuficientes = []
            for detalle in detalles:
                prod = prod_map.get(detalle.producto.id) if detalle.producto else None
                if not prod:
                    insuficientes.append((detalle, "Producto no encontrado"))
                    continue
                try:
                    stock_actual = int(prod.stock or 0)
                except Exception:
                    stock_actual = 0
                if detalle.cantidad > stock_actual:
                    insuficientes.append((detalle, f"Solicitado {detalle.cantidad}, Disponible {stock_actual}"))

            if insuficientes:
                msg = " No hay suficiente stock:\n" + "\n".join(
                    f"{it.producto.nombre if it.producto else '??'}: {reason}" for it, reason in insuficientes
                )
                print("[confirmar_pedido] insuficientes:", insuficientes)
                messages.error(request, msg)
                return redirect('ventas:detalle_pedido', pedido_id=pedido.id)

            # 2) Aplicar decrementos con F() y verificar filas afectadas
            for detalle in detalles:
                prod = prod_map.get(detalle.producto.id)
                descuento = int(detalle.cantidad)
                print(f"[confirmar_pedido] intentanto decrementar producto {prod.id} -{descuento} (stock antes: {prod.stock})")

                # Realizamos update atómico: solo se decrementa si stock >= descuento
                rows = Producto.objects.filter(id=prod.id, stock__gte=descuento).update(stock=F('stock') - descuento)
                if rows == 0:
                    raise RuntimeError(f"No se pudo decrementar stock para producto {prod.id} (rows affected=0)")

                # Para depuración: obtener nuevo valor
                nuevo = Producto.objects.get(id=prod.id).stock
                print(f"[confirmar_pedido] producto {prod.id} stock actualizado -> {nuevo}")

            # 3) Actualizar pedido y documento
            pedido.estado = 'Procesando'
            pedido.save()

            if hasattr(pedido, 'documentoventa') and pedido.documentoventa:
                doc = pedido.documentoventa
                if doc.estado != 'Emitida':
                    doc.estado = 'Pagada'
                    doc.save()

            messages.success(request, f'El Pedido #{pedido.id} ha sido confirmado y el stock descontado.')
            print(f"[confirmar_pedido] OK pedido {pedido.id} confirmado y stock actualizado.")

    except Exception as e:
        print("[confirmar_pedido] ERROR:", repr(e))
        messages.error(request, f'Ocurrió un error al confirmar el pedido: {e}')
        return redirect('ventas:detalle_pedido', pedido_id=pedido.id)

    return redirect('ventas:detalle_pedido', pedido_id=pedido.id)

@login_required
def marcar_pedido_enviado(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)

    if pedido.estado != 'Procesando':
        messages.warning(request, 'Solo se pueden marcar como enviados los pedidos en proceso.')
        return redirect('ventas:detalle_pedido', pedido_id=pedido.id)

    pedido.estado = 'Enviado'
    pedido.save()

    messages.success(request, f'El pedido #{pedido.id} ha sido marcado como Enviado.')
    return redirect('ventas:detalle_pedido', pedido_id=pedido.id)



# REPORTES Y ESTADÍSTICAS


@login_required
def estadisticas_ventas(request):
    if request.user.rol not in ['Administrador', 'Tesoreria']:
        messages.error(request, " No tienes permisos para acceder a esta sección.")
        return redirect('usuarios:dashboard')

    pedidos = (Pedido.objects.filter(estado='Enviado')
               .select_related('cliente', 'usuario', 'documentoventa')
               .order_by('-fecha_creacion'))

    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    vendedor_query = request.GET.get('vendedor') 

    # Aplicar filtros de fecha
    if fecha_desde:
        pedidos = pedidos.filter(fecha_creacion__date__gte=fecha_desde)
    if fecha_hasta:
        pedidos = pedidos.filter(fecha_creacion__date__lte=fecha_hasta)

    # Aplicar filtro por vendedor
    if vendedor_query:
        pedidos = pedidos.filter(
            Q(usuario__username__icontains=vendedor_query) | 
            Q(usuario__first_name__icontains=vendedor_query) |
            Q(usuario__last_name__icontains=vendedor_query)
        )

    # CÁLCULO DE TOTALES (se mantiene igual)
    total_ventas = pedidos.count()
    monto_total = sum(
        p.documentoventa.total 
        for p in pedidos 
        if hasattr(p, 'documentoventa') and p.documentoventa and p.documentoventa.total
    )


    TASA_COMISION = Decimal('0.01')
    comision_total = monto_total * TASA_COMISION

    context = {
        'pedidos': pedidos,
        'total_ventas': total_ventas,
        'monto_total': monto_total,
        'comision_total': comision_total, 
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'vendedor': vendedor_query,
    }

    return render(request, 'ventas/estadisticas_ventas.html', context)

@login_required
def exportar_ventas_excel(request):
    if request.user.rol not in ['Administrador', 'Tesoreria']:
        messages.error(request, " No tienes permisos para exportar.")
        return redirect('usuarios:dashboard')

    pedidos = (Pedido.objects.filter(estado='Enviado')
               .select_related('cliente', 'usuario', 'documentoventa')
               .order_by('-fecha_creacion'))

    fecha_desde_str = request.GET.get('fecha_desde')
    fecha_hasta_str = request.GET.get('fecha_hasta')

    if fecha_desde_str:
        try:
            fecha_desde = datetime.strptime(fecha_desde_str, "%Y-%m-%d").date()
            pedidos = pedidos.filter(fecha_creacion__date__gte=fecha_desde)
        except ValueError:
            messages.error(request, " Fecha desde inválida. Usa formato YYYY-MM-DD.")

    if fecha_hasta_str:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_str, "%Y-%m-%d").date()
            pedidos = pedidos.filter(fecha_creacion__date__lte=fecha_hasta)
        except ValueError:
            messages.error(request, " Fecha hasta inválida. Usa formato YYYY-MM-DD.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = ['Pedido #', 'Cliente', 'RUT', 'Vendedor', 'Fecha', 'Estado', 'Total']
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for pedido in pedidos:
        total = pedido.documentoventa.total if hasattr(pedido, 'documentoventa') and pedido.documentoventa else 0
        total_str = f"${total:,.0f}".replace(",", ".") 

        ws.append([
            pedido.id,
            pedido.cliente.razon_social,
            pedido.cliente.rut,
            pedido.usuario.get_full_name() or pedido.usuario.username,
            pedido.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
            pedido.estado,
            total_str 
        ])

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"ventas_{date.today().strftime('%d-%m-%Y')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
def vista_checkout(request):
    try:
        cliente_actual = request.user.perfil_cliente
    except Cliente.DoesNotExist:
        return redirect('clientes:completar_perfil') 
    except AttributeError:
        return redirect('usuarios:dashboard') 

    if request.method == 'POST':
        form = CheckoutForm(request.POST, instance=cliente_actual)
        if form.is_valid():
            return redirect('ventas:compra_exitosa') 
    else:
        form = CheckoutForm(instance=cliente_actual)
    context = { 'form': form }
    return render(request, 'ventas/checkout.html', context)

@login_required
def exportar_reporte_rentabilidad(request):
    """
    Genera el Excel de Reporte de Ventas mensual con cálculo de utilidad.
    """
    if request.user.rol not in ['Administrador', 'Tesoreria']:
        messages.error(request, "⚠️ No tienes permisos para exportar este reporte.")
        return redirect('usuarios:dashboard')

    # 1. Obtener los detalles de documentos 
    detalles_vendidos = DetalleDocumento.objects.filter(
        documento__pedido__estado='Enviado' 
    ).select_related(
        'documento__vendedor', 'documento__cliente', 
        'producto', 'producto__proveedor'
    ).order_by('-documento__fecha_emision')

    # 2. Aplicar filtros de fecha 
    fecha_desde_str = request.GET.get('fecha_desde')
    fecha_hasta_str = request.GET.get('fecha_hasta')

    # Filtro Fecha Desde
    if fecha_desde_str and fecha_desde_str.strip(): 
        try:
            fecha_desde = datetime.strptime(fecha_desde_str.strip(), "%Y-%m-%d").date()
            detalles_vendidos = detalles_vendidos.filter(documento__fecha_emision__date__gte=fecha_desde)
        except ValueError:
            messages.error(request, "⚠️ Error en la fecha de inicio. Usa el formato YYYY-MM-DD.")

    # Filtro Fecha Hasta
    if fecha_hasta_str and fecha_hasta_str.strip(): 
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_str.strip(), "%Y-%m-%d").date()
            detalles_vendidos = detalles_vendidos.filter(documento__fecha_emision__date__lte=fecha_hasta)
        except ValueError:
            messages.error(request, "⚠️ Error en la fecha de fin. Usa el formato YYYY-MM-DD.")

    # 3. Crear el Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Rentabilidad"

    # Estilos de encabezado
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    headers = [
        'Fecha Venta', 'Documento', 'Folio', 'Vendedor', 'Cliente', 
        'Proveedor', 'Producto (SKU)', 'Cantidad', 
        'Valor Costo (Unit.)', 'Valor Venta (Unit. Neto)', 
        'Costo Total', 'Venta Total (Neta)', 'Utilidad', 'Margen (%)'
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 4. Llenar el Excel con los datos
    for detalle in detalles_vendidos:
        
        # --- Obtener Datos ---
        documento = detalle.documento
        producto = detalle.producto
        cantidad = detalle.cantidad

        vendedor_nombre = documento.vendedor.username if documento.vendedor else 'N/A'
        cliente_nombre = documento.cliente.razon_social
        proveedor_nombre = producto.proveedor.razon_social if producto.proveedor else 'N/A'
        
        # --- Cálculos de Rentabilidad (Lógica de IVA Inversa) ---
        precio_venta_bruto_unit = detalle.precio_unitario_venta
        precio_venta_neto_unit = (precio_venta_bruto_unit / Decimal('1.19')).quantize(Decimal('0.00'))
        
        costo_unit = detalle.costo_unitario_venta or producto.costo_unitario or Decimal('0')
        
        # Totales por línea
        costo_total_linea = costo_unit * cantidad
        venta_neta_total_linea = precio_venta_neto_unit * cantidad
        
        # Utilidad y Margen
        utilidad_linea = venta_neta_total_linea - costo_total_linea
        margen_linea = 0
        if venta_neta_total_linea > 0:
            margen_linea = (utilidad_linea / venta_neta_total_linea) * 100

        # --- Escribir Fila ---
        ws.append([
            documento.fecha_emision.strftime('%d/%m/%Y'),
            documento.tipo_documento,
            documento.folio,
            vendedor_nombre,
            cliente_nombre,
            proveedor_nombre,
            producto.nombre, 
            cantidad,
            costo_unit,
            precio_venta_neto_unit,
            costo_total_linea,
            venta_neta_total_linea,
            utilidad_linea,
            f"{margen_linea:.2f}%"
        ])


    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"reporte_rentabilidad_{date.today().strftime('%d-%m-%Y')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response